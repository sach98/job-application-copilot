#!/usr/bin/env python3
import sys
from pathlib import Path
from openpyxl import Workbook

sys.path.append(str(Path(__file__).parent))
from lib.paths import PROFILE_DIR

OUTPUT_PATH = PROFILE_DIR / "JobHunt_Tracker.xlsx"

def main() -> int:
    print(f"[*] Creating Excel sheet template at {OUTPUT_PATH}...", file=sys.stderr)
    
    wb = Workbook()
    
    # --- Tab 1: applications ---
    ws1 = wb.active
    ws1.title = "applications"
    headers_applications = [
        "id", "date", "company", "role", "source", "jd_url", "posted_at", "salary", 
        "fit_score", "score_components", "hiring_mgr_name", "hiring_mgr_title", 
        "hiring_mgr_linkedin", "hiring_mgr_email", "team_members", "referrals", 
        "dossier_url", "resume_url", "cl_url", "essay_answers_url", "screenshots_url",
        "cl_preview", "resume_diff_preview", "fit_summary",
        "status", "queued_at", "applied_at", "followup_1_due", "followup_1_sent_at",
        "followup_2_due", "followup_2_sent_at", "response_at", "response_summary", 
        "interview_date", "offer_details", "notes", "referral_asked"
    ]
    ws1.append(headers_applications)
    
    # --- Tab 2: daily_cap_tracker ---
    ws2 = wb.create_sheet(title="daily_cap_tracker")
    headers_cap = ["date", "cap", "applied_count", "pending_review_count", "skipped_count"]
    ws2.append(headers_cap)
    
    # --- Tab 3: weekly_review ---
    ws3 = wb.create_sheet(title="weekly_review")
    headers_weekly = [
        "week_ending", "applications_sent", "responses_received", "response_rate", 
        "interviews_scheduled", "top_source", "avg_fit_score_applied", "avg_fit_score_responded"
    ]
    ws3.append(headers_weekly)
    
    # --- Tab 4: enrichment_cache ---
    ws4 = wb.create_sheet(title="enrichment_cache")
    headers_enrichment = ["company", "domain", "email_pattern", "linkedin_company_url", "last_refreshed"]
    ws4.append(headers_enrichment)
    
    # Save Workbook
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    
    print(f"[+] Successfully generated template: {OUTPUT_PATH}", file=sys.stderr)
    return 0

if __name__ == "__main__":
    sys.exit(main())
